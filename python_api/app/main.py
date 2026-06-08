from __future__ import annotations

import csv
from datetime import datetime
import json
import logging
import os
from pathlib import Path
import re
import secrets
import time
import unicodedata
from typing import Optional

from fastapi import Depends, FastAPI, Header, HTTPException, Query, Request, status
from fastapi.middleware.cors import CORSMiddleware
from passlib.context import CryptContext
from sqlalchemy import distinct, func, select
from sqlalchemy.orm import Session

from .database import Base, SessionLocal, engine
from .models import Answer, Assessment, AuthToken, Company, Question, User
from .schemas import (
    ForgotPasswordRequest,
    LoginRequest,
    ResetPasswordRequest,
    SaveAssessmentRequest,
    SignupCompanyRequest,
    UpdateUserProfileRequest,
)


app = FastAPI(title="Bot Soberania Python API", version="0.2.0")
pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
logger = logging.getLogger("bot_soberania_api")

AUTO_CREATE_SCHEMA = os.getenv("AUTO_CREATE_SCHEMA", "1") == "1"
SKIP_STARTUP_BACKFILL = os.getenv("SKIP_STARTUP_BACKFILL", "0") == "1"
TOKEN_TTL_MINUTES = int(os.getenv("TOKEN_TTL_MINUTES", "1440"))  # 24h
PASSWORD_MIN_LENGTH = int(os.getenv("PASSWORD_MIN_LENGTH", "8"))
RATE_LIMIT_WINDOW_SECONDS = int(os.getenv("RATE_LIMIT_WINDOW_SECONDS", "60"))
RATE_LIMIT_MAX_ATTEMPTS = int(os.getenv("RATE_LIMIT_MAX_ATTEMPTS", "15"))
RATE_LIMITED_PATHS = {"/login", "/forgot_password", "/reset_password", "/signup_company"}
ALLOW_DEV_BOOTSTRAP_LOGIN = os.getenv("ALLOW_DEV_BOOTSTRAP_LOGIN", "1") == "1"
ALLOW_LEGACY_PASSWORD_LOGIN = os.getenv("ALLOW_LEGACY_PASSWORD_LOGIN", "1") == "1"
ALLOWED_ORIGIN_REGEX = os.getenv(
    "ALLOWED_ORIGIN_REGEX",
    # Local + Firebase Hosting (*.web.app / *.firebaseapp.com).
    r"^https?://(\[::1\](:\d+)?|(localhost|127\.0\.0\.1)(:\d+)?)$"
    r"|^https://[a-zA-Z0-9.-]+\.web\.app$"
    r"|^https://[a-zA-Z0-9.-]+\.firebaseapp\.com$",
)

# key: "<ip>:<path>" -> [timestamps]
RATE_LIMIT_BUCKETS: dict[str, list[float]] = {}


def _norm_text(value: object) -> str:
    s = str(value or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s


def _safe_parent(base: Path, level: int) -> Optional[Path]:
    try:
        return base.parents[level]
    except IndexError:
        return None


def _data_file_candidates(filename: str) -> list[Path]:
    here = Path(__file__).resolve()
    candidates: list[Path] = []
    for base in (
        _safe_parent(here, 1),  # python_api/ (Docker: /app)
        _safe_parent(here, 2),  # repo root em dev local
        _safe_parent(here, 3),
    ):
        if base is None:
            continue
        candidates.append(base / "data" / filename)
        candidates.append(base / filename)
    # dedupe preservando ordem
    seen: set[str] = set()
    unique: list[Path] = []
    for path in candidates:
        key = str(path)
        if key not in seen:
            seen.add(key)
            unique.append(path)
    return unique


def _resolve_data_file(filename: str, env_var: str = "") -> Optional[Path]:
    override = os.getenv(env_var, "").strip() if env_var else ""
    if override:
        path = Path(override)
        return path if path.is_file() else None
    for candidate in _data_file_candidates(filename):
        if candidate.is_file():
            return candidate
    return None


def _load_technical_pilar_maps() -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    xlsx_name = "Assessment_OTIMIZADO_176_Perguntas_v4 1 (2).xlsx"
    xlsx_candidates = [
        os.getenv("DEFAULT_TECH_PILAR_XLSX", "").strip(),
        *[
            str(path)
            for path in _data_file_candidates(xlsx_name)
        ],
        r"C:\Users\user\Downloads\Assessment_OTIMIZADO_176_Perguntas_v4 1 (2).xlsx",
    ]
    xlsx_path = next((p for p in xlsx_candidates if p and Path(p).is_file()), "")
    if not xlsx_path:
        logger.warning("Arquivo XLSX de pilares tecnicos nao encontrado.")
        return {}, {}, {}

    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl nao instalado; backfill de pilar_tecnico ignorado.")
        return {}, {}, {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Assessment"] if "Assessment" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_idx = None
    header_vals: list[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), start=1):
        vals = [str(c).strip() if c is not None else "" for c in row]
        norm_vals = [_norm_text(v) for v in vals]
        if any(("pilar" in v and "tecn" in v and "aws" in v) for v in norm_vals):
            header_idx = i
            header_vals = vals
            break

    if header_idx is None:
        wb.close()
        logger.warning("Cabecalho do XLSX com pilar tecnico nao encontrado: %s", xlsx_path)
        return {}, {}, {}

    def _find_col(pred) -> Optional[int]:
        for idx, v in enumerate(header_vals):
            if pred(_norm_text(v)):
                return idx
        return None

    col_code = _find_col(lambda v: "codigo" in v)
    col_question = _find_col(lambda v: "pergunta" in v)
    col_tech = _find_col(lambda v: ("pilar" in v and "tecn" in v and "aws" in v))
    col_visual = _find_col(lambda v: ("visualizar" in v and "soberania" in v))
    col_domain = _find_col(lambda v: ("dominio" in v and "soberania" in v))
    if col_tech is None:
        wb.close()
        return {}, {}, {}

    by_code: dict[str, str] = {}
    by_recommendation: dict[str, str] = {}
    domain_counts: dict[str, dict[str, int]] = {}
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if col_visual is not None:
            visual = vals[col_visual] if col_visual < len(vals) else ""
            if _norm_text(visual) not in {"ok"}:
                continue
        tech = vals[col_tech] if col_tech < len(vals) else ""
        if not tech:
            continue
        domain = ""
        if col_domain is not None and col_domain < len(vals):
            domain = vals[col_domain].strip()
        if col_code is not None and col_code < len(vals):
            code = vals[col_code].strip()
            if code and code not in by_code:
                by_code[code] = tech
        if col_question is not None and col_question < len(vals):
            question = vals[col_question].strip()
            if question and question not in by_recommendation:
                by_recommendation[question] = tech
        if domain:
            dk = _norm_text(domain)
            bucket = domain_counts.setdefault(dk, {})
            bucket[tech] = int(bucket.get(tech, 0) + 1)

    wb.close()
    by_domain: dict[str, str] = {}
    for dk, counts in domain_counts.items():
        if not counts:
            continue
        best = sorted(counts.items(), key=lambda x: x[1], reverse=True)[0][0]
        by_domain[dk] = best
    return by_code, by_recommendation, by_domain


def _backfill_question_technical_pilar(db: Session) -> int:
    by_code, by_recommendation, by_domain = _load_technical_pilar_maps()
    if not by_code and not by_recommendation and not by_domain:
        return 0

    updated = 0
    questions = db.scalars(select(Question)).all()
    for q in questions:
        if (q.pilar_tecnico or "").strip():
            continue
        tech = None
        code = (q.question_code or "").strip()
        if code:
            tech = by_code.get(code)
        if not tech:
            tech = by_recommendation.get((q.recommendation or "").strip())
        if not tech:
            tech = by_domain.get(_norm_text(q.dominio or ""))
        if tech:
            q.pilar_tecnico = tech
            db.add(q)
            updated += 1
    if updated:
        db.commit()
    return updated


def _load_domain_map_from_csv() -> tuple[dict[str, str], dict[str, str]]:
    csv_path = _resolve_data_file("Data (1).csv", "DEFAULT_DOMAINS_CSV")
    if csv_path is None:
        logger.warning("Arquivo de domínio não encontrado.")
        return {}, {}

    by_code: dict[str, str] = {}
    by_recommendation: dict[str, str] = {}
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            domain = (row.get("Dominio") or "").strip()
            if not domain:
                continue
            code = (row.get("Unnamed: 0") or "").strip()
            recommendation = (row.get("Recommendation") or "").strip()
            if code and code not in by_code:
                by_code[code] = domain
            if recommendation and recommendation not in by_recommendation:
                by_recommendation[recommendation] = domain
    return by_code, by_recommendation


def _backfill_question_domains(db: Session) -> int:
    by_code, by_recommendation = _load_domain_map_from_csv()
    if not by_code and not by_recommendation:
        return 0

    updated = 0
    questions = db.scalars(select(Question)).all()
    for q in questions:
        if (q.dominio or "").strip():
            continue
        domain = None
        code = (q.question_code or "").strip()
        if code:
            domain = by_code.get(code)
        if not domain:
            domain = by_recommendation.get((q.recommendation or "").strip())
        if domain:
            q.dominio = domain
            db.add(q)
            updated += 1
    if updated:
        db.commit()
    return updated


def _load_default_questions(db: Session) -> int:
    csv_path = _resolve_data_file(
        "Prototipo_Assessment_Tool_Maturidade_Soberania_v1_with_order_index.csv",
        "DEFAULT_QUESTIONS_CSV",
    )
    if csv_path is None:
        logger.warning("Arquivo de perguntas padrão não encontrado.")
        return 0

    rows: list[Question] = []
    encodings = ("utf-8-sig", "cp1252", "latin-1")
    for enc in encodings:
        try:
            with csv_path.open("r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f, delimiter=";")
                for i, row in enumerate(reader, start=1):
                    recommendation = (row.get("Recommendation") or "").strip()
                    phase = (row.get("Phase") or "").strip()
                    pilar = (row.get("Pilares") or "").strip()
                    if not recommendation or not phase or not pilar:
                        continue
                    rows.append(
                        Question(
                            created_at=datetime.utcnow(),
                            phase=phase,
                            pilar=pilar,
                            recommendation=recommendation,
                            weight=None,
                            order_index=i,
                            question_code=(row.get("Unnamed: 0") or "").strip() or None,
                            dominio=None,
                            recommendation_en=None,
                            recommendation_es=None,
                            guidance=(row.get("Guidance for assessments") or "").strip() or None,
                            how_to_check=(row.get("How to check") or "").strip() or None,
                            aws_service=(row.get("Associated \nAWS Service") or "").strip() or None,
                            pilar_tecnico=None,
                            norms=" · ".join(
                                [
                                    x
                                    for x in (
                                        (row.get("Normas Técnicas") or "").strip(),
                                        (row.get("Geral (Brasil)") or "").strip(),
                                        (row.get("Finanças Públicas (BCB)") or "").strip(),
                                        (row.get("Educação") or "").strip(),
                                    )
                                    if x
                                ]
                            )
                            or None,
                        )
                    )
            break
        except UnicodeDecodeError:
            rows = []
            continue

    if not rows:
        logger.warning("Nenhuma pergunta válida encontrada no CSV padrão: %s", csv_path)
        return 0

    db.add_all(rows)
    db.commit()
    return len(rows)


app.add_middleware(
    CORSMiddleware,
    allow_origins=[],
    allow_origin_regex=ALLOWED_ORIGIN_REGEX,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup() -> None:
    # Mantem compatibilidade local. Em ambientes com Alembic, use AUTO_CREATE_SCHEMA=0.
    if AUTO_CREATE_SCHEMA:
        Base.metadata.create_all(bind=engine)
    if SKIP_STARTUP_BACKFILL:
        logger.info("SKIP_STARTUP_BACKFILL=1: bootstrap/backfill de arquivos locais ignorado.")
        return
    try:
        with SessionLocal() as db:
            has_questions = int(db.scalar(select(func.count(Question.id))) or 0)
            if has_questions == 0:
                inserted = _load_default_questions(db)
                if inserted:
                    logger.info("Bootstrap local: %s perguntas importadas automaticamente.", inserted)
            updated_domains = _backfill_question_domains(db)
            if updated_domains:
                logger.info("Backfill local: %s perguntas atualizadas com domínio.", updated_domains)
            updated_tech = _backfill_question_technical_pilar(db)
            if updated_tech:
                logger.info("Backfill local: %s perguntas atualizadas com pilar técnico.", updated_tech)
    except Exception:
        logger.exception("Bootstrap/backfill no startup falhou; API continua online.")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _validate_password_policy(password: str) -> None:
    if len(password) < PASSWORD_MIN_LENGTH:
        raise HTTPException(
            status_code=400,
            detail=f"Password must have at least {PASSWORD_MIN_LENGTH} characters",
        )
    if not re.search(r"[A-Za-z]", password) or not re.search(r"\d", password):
        raise HTTPException(
            status_code=400,
            detail="Password must contain at least one letter and one number",
        )


def _apply_rate_limit(request: Request) -> None:
    path = request.url.path
    if path not in RATE_LIMITED_PATHS:
        return
    ip = request.client.host if request.client else "unknown"
    key = f"{ip}:{path}"
    now = time.time()
    window_start = now - RATE_LIMIT_WINDOW_SECONDS
    timestamps = [t for t in RATE_LIMIT_BUCKETS.get(key, []) if t >= window_start]
    if len(timestamps) >= RATE_LIMIT_MAX_ATTEMPTS:
        raise HTTPException(
            status_code=429,
            detail="Too many attempts. Please try again later.",
        )
    timestamps.append(now)
    RATE_LIMIT_BUCKETS[key] = timestamps


def _generate_token() -> str:
    exp_ts = int(time.time() + TOKEN_TTL_MINUTES * 60)
    nonce = secrets.token_urlsafe(24)
    return f"v1.{exp_ts}.{nonce}"


def _token_is_expired(token_value: str) -> bool:
    parts = token_value.split(".")
    if len(parts) < 3 or parts[0] != "v1":
        return True
    try:
        exp_ts = int(parts[1])
    except ValueError:
        return True
    return time.time() > exp_ts


def _extract_bearer(authorization: Optional[str]) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    prefix = "Bearer "
    if not authorization.startswith(prefix):
        raise HTTPException(status_code=401, detail="Invalid Authorization header format")
    token = authorization[len(prefix) :].strip()
    if not token:
        raise HTTPException(status_code=401, detail="Invalid Authorization token")
    return token


def get_current_user(
    authorization: Optional[str] = Header(default=None),
    db: Session = Depends(get_db),
) -> User:
    token_value = _extract_bearer(authorization)
    if _token_is_expired(token_value):
        # limpeza defensiva do token expirado
        token_obj = db.get(AuthToken, token_value)
        if token_obj:
            db.delete(token_obj)
            db.commit()
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    token = db.get(AuthToken, token_value)
    if not token:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    user = db.get(User, token.user_id)
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


@app.get("/health")
def health():
    return {
        "ok": True,
        "service": "python_api",
        "db": "connected",
        "security": {
            "token_ttl_minutes": TOKEN_TTL_MINUTES,
            "password_min_length": PASSWORD_MIN_LENGTH,
            "rate_limit_window_seconds": RATE_LIMIT_WINDOW_SECONDS,
            "rate_limit_max_attempts": RATE_LIMIT_MAX_ATTEMPTS,
        },
    }


@app.post("/forgot_password")
def forgot_password(
    payload: ForgotPasswordRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    _apply_rate_limit(request)
    user = db.scalar(select(User).where(User.email == payload.email.lower()))
    if user:
        reset = {"token": secrets.token_hex(16), "expiration": None, "used": False}
        user.password_reset = json.dumps(reset, ensure_ascii=False)
        db.add(user)
        db.commit()
    return {"ok": True, "email": payload.email}


@app.post("/reset_password")
def reset_password(
    payload: ResetPasswordRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    _apply_rate_limit(request)
    _validate_password_policy(payload.new_password)
    user = db.scalar(select(User).where(User.email == payload.email.lower()))
    if user:
        user.password_hash = pwd_context.hash(payload.new_password)
        user.legacy_password_hash = None
        user.password_reset = json.dumps(
            {"token": payload.token, "expiration": None, "used": True},
            ensure_ascii=False,
        )
        db.add(user)
        db.commit()
    return {"ok": True}


@app.post("/login")
def login(payload: LoginRequest, request: Request, db: Session = Depends(get_db)):
    _apply_rate_limit(request)
    email = payload.email.lower()
    user = db.scalar(select(User).where(User.email == email))
    if not user:
        # Bootstrap local: em ambiente de desenvolvimento, se nao houver usuarios
        # ainda, cria o primeiro usuario a partir do proprio login.
        if ALLOW_DEV_BOOTSTRAP_LOGIN and AUTO_CREATE_SCHEMA:
            total_users = int(db.scalar(select(func.count(User.id))) or 0)
            if total_users == 0:
                _validate_password_policy(payload.password)
                display_name = email.split("@")[0]
                user = User(
                    name=display_name,
                    email=email,
                    password_hash=pwd_context.hash(payload.password),
                    role="admin",
                    created_at=datetime.utcnow(),
                )
                db.add(user)
                db.flush()
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")

    valid = False
    if user.password_hash:
        valid = pwd_context.verify(payload.password, user.password_hash)
    elif ALLOW_LEGACY_PASSWORD_LOGIN and user.legacy_password_hash:
        legacy = (user.legacy_password_hash or "").strip()
        if legacy:
            if legacy.startswith("$"):
                try:
                    valid = pwd_context.verify(payload.password, legacy)
                except (ValueError, TypeError):
                    valid = False
            else:
                valid = secrets.compare_digest(payload.password, legacy)
            if valid:
                user.password_hash = pwd_context.hash(payload.password)
                user.legacy_password_hash = None
                db.add(user)

    if not valid:
        logger.warning("Invalid login attempt for email=%s", email)
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = _generate_token()
    db.add(AuthToken(token=token, user_id=user.id))
    db.commit()

    display_name = user.name or email.split("@")[0]
    return {
        "authToken": token,
        "user": {
            "id": user.id,
            "name": display_name,
            "email": user.email,
            "role": user.role or "user",
        },
        "admin_name": display_name,
        "name": display_name,
        "role": user.role or "user",
    }


@app.post("/signup_company")
def signup_company(
    payload: SignupCompanyRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    _apply_rate_limit(request)
    _validate_password_policy(payload.password)
    email = payload.email.lower()
    existing = db.scalar(select(User).where(User.email == email))
    if existing:
        raise HTTPException(status_code=409, detail="Email already registered")

    company = Company(
        name=payload.company_name.strip(),
        cnpj=payload.cnpj,
        segment=payload.segment,
        status="ACTIVE",
        created_at=datetime.utcnow(),
    )
    db.add(company)
    db.flush()

    user = User(
        name=payload.name.strip(),
        email=email,
        password_hash=pwd_context.hash(payload.password),
        role=payload.role,
        company_id=company.id,
        last_name=payload.last_name,
        phone=payload.phone,
        created_at=datetime.utcnow(),
    )
    db.add(user)
    db.flush()

    token = _generate_token()
    db.add(AuthToken(token=token, user_id=user.id))
    db.commit()

    return {
        "authToken": token,
        "user": {"id": user.id, "name": user.name, "email": user.email},
        "admin_name": payload.admin_name,
        "name": user.name,
    }


@app.post("/assessment/resume")
def assessment_resume(_: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    assessment = db.scalar(
        select(Assessment)
        .where(Assessment.created_by == current_user.id)
        .order_by(Assessment.id.desc())
    )
    if not assessment:
        assessment = Assessment(
            company_id=current_user.company_id,
            status="IN_PROGRESS",
            current_phase="QUICK_WINS",
            last_question=0,
            progress_percent=0,
            created_by=current_user.id,
            created_at=datetime.utcnow(),
        )
        db.add(assessment)
        db.commit()
        db.refresh(assessment)
    return {"id": assessment.id}


@app.get("/questions")
def list_questions(
    phase: str = Query(...),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    rows = db.scalars(select(Question).where(Question.phase == phase).order_by(Question.order_index.asc())).all()
    return [
        {
            "id": q.id,
            "phase": q.phase,
            "pilar": q.pilar,
            "dominio": q.dominio,
            "recommendation": q.recommendation,
            "recommendation_en": q.recommendation_en,
            "recommendation_es": q.recommendation_es,
            "guidance": q.guidance,
            "how_to_check": q.how_to_check,
            "order_index": q.order_index or 0,
            "question_code": q.question_code,
            "aws_service": q.aws_service,
            "pilar_tecnico": q.pilar_tecnico,
        }
        for q in rows
    ]


@app.get("/progress/assessment")
def get_progress(
    assessment_id: int = Query(...),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    rows = db.scalars(
        select(Answer).where(Answer.assessment_id == assessment_id).order_by(Answer.id.asc())
    ).all()
    return {
        "assessment_id": assessment_id,
        "answers": [
            {
                "id": a.id,
                "question": a.question_id,
                "score": a.score,
                "justification": a.justification,
                "evidence": a.evidence,
            }
            for a in rows
        ],
    }


def _require_admin(current_user: User) -> User:
    if (current_user.role or "").lower() != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


_ADMIN_PROGRESS_TEXT_MAX = 768


def _truncate_admin_field(value: Optional[str], max_len: int = _ADMIN_PROGRESS_TEXT_MAX) -> Optional[str]:
    """Limita tamanho de Text no JSON admin (evita respostas de vários MB no browser)."""
    if value is None:
        return None
    s = str(value)
    if len(s) <= max_len:
        return s
    return s[: max_len - 3] + "..."


def _admin_answer_count(db: Session, assessment_id: int) -> int:
    n = db.scalar(
        select(func.count(distinct(Answer.question_id))).where(
            Answer.assessment_id == assessment_id
        )
    )
    return int(n or 0)


def _question_catalog_stats(db: Session) -> tuple[int, dict[str, int]]:
    total = int(db.scalar(select(func.count(Question.id))) or 0)
    rows = db.execute(
        select(Question.pilar, func.count(Question.id)).group_by(Question.pilar)
    ).all()
    by_pilar: dict[str, int] = {}
    for pilar, count in rows:
        key = (pilar or "").strip()
        if key:
            by_pilar[key] = int(count or 0)
    return total, by_pilar


def _admin_answers_payload(
    db: Session,
    assessment_id: int,
    *,
    offset: int = 0,
    limit: Optional[int] = None,
) -> list[dict]:
    stmt = (
        select(Answer, Question)
        .select_from(Answer)
        .outerjoin(Question, Answer.question_id == Question.id)
        .where(Answer.assessment_id == assessment_id)
        .order_by(Answer.id.asc())
    )
    if limit is not None:
        stmt = stmt.offset(offset).limit(limit)
    rows = db.execute(stmt).all()
    answer_details: list[dict] = []
    for a, q in rows:
        answer_details.append(
            {
                "question_id": a.question_id,
                "question_code": q.question_code if q else None,
                "pilar": q.pilar if q else None,
                "dominio": q.dominio if q else None,
                "pilar_tecnico": q.pilar_tecnico if q else None,
                "recommendation": _truncate_admin_field(q.recommendation if q else None),
                "aws_service": _truncate_admin_field(q.aws_service if q else None),
                "norms": _truncate_admin_field(q.norms if q else None),
                "score": a.score,
                "justification": _truncate_admin_field(a.justification),
            }
        )
    return answer_details


@app.get("/admin/users")
def admin_list_users(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    total_questions, questions_by_pilar = _question_catalog_stats(db)
    users = db.scalars(select(User).order_by(User.created_at.desc())).all()
    result = []
    for u in users:
        company = db.get(Company, u.company_id) if u.company_id else None
        assessment = db.scalar(
            select(Assessment)
            .where(Assessment.created_by == u.id)
            .order_by(Assessment.id.desc())
        )
        answered_count = 0
        if assessment:
            answered_count = int(
                db.scalar(
                    select(func.count(distinct(Answer.question_id))).where(
                        Answer.assessment_id == assessment.id
                    )
                )
                or 0
            )
        result.append({
            "id": u.id,
            "name": u.name,
            "last_name": u.last_name,
            "email": u.email,
            "role": u.role or "user",
            "phone": u.phone,
            "created_at": u.created_at.isoformat() if u.created_at else None,
            "company": {
                "id": company.id,
                "name": company.name,
                "cnpj": company.cnpj,
                "segment": company.segment,
            } if company else None,
            "assessment": {
                "id": assessment.id,
                "status": assessment.status,
                "current_phase": assessment.current_phase,
                "progress_percent": assessment.progress_percent,
                "answered_count": answered_count,
                "total_questions": total_questions,
                "questions_by_pilar": questions_by_pilar,
            } if assessment else None,
        })
    return result


@app.get("/admin/users/{user_id}/progress")
def admin_user_progress(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Retorna perfil + lista de assessments (sem `answers`).

    O payload de respostas quebra o browser (Failed to fetch) em usuarios com
    muitas respostas. Sempre use GET /admin/assessments/{id}/answers apos carregar
    esta rota (o app Flutter ja faz isso).
    """
    _require_admin(current_user)
    total_questions, questions_by_pilar = _question_catalog_stats(db)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    assessments = db.scalars(
        select(Assessment)
        .where(Assessment.created_by == user_id)
        .order_by(Assessment.id.desc())
    ).all()

    result = []
    for assessment in assessments:
        cnt = _admin_answer_count(db, assessment.id)
        result.append(
            {
                "id": assessment.id,
                "status": assessment.status,
                "current_phase": assessment.current_phase,
                "progress_percent": assessment.progress_percent,
                "created_at": assessment.created_at.isoformat()
                if assessment.created_at
                else None,
                "answers": [],
                "answer_count": cnt,
                "total_questions": total_questions,
                "questions_by_pilar": questions_by_pilar,
            }
        )

    company = db.get(Company, user.company_id) if user.company_id else None
    return {
        "user": {
            "id": user.id,
            "name": user.name,
            "last_name": user.last_name,
            "email": user.email,
            "role": user.role or "user",
            "phone": user.phone,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "company": {
                "id": company.id,
                "name": company.name,
                "cnpj": company.cnpj,
                "segment": company.segment,
                "status": company.status,
            } if company else None,
        },
        "assessments": result,
    }


@app.get("/admin/assessments/{assessment_id}/answers")
def admin_assessment_answers(
    assessment_id: int,
    offset: int = Query(0, ge=0),
    limit: int = Query(60, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    if not db.get(Assessment, assessment_id):
        raise HTTPException(status_code=404, detail="Assessment not found")
    total = _admin_answer_count(db, assessment_id)
    answer_details = _admin_answers_payload(
        db, assessment_id, offset=offset, limit=limit
    )
    return {
        "answers": answer_details,
        "total": total,
        "offset": offset,
        "limit": limit,
    }


def _clean_optional_str(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    stripped = value.strip()
    return stripped if stripped else None


@app.put("/admin/users/{user_id}")
def admin_update_user(
    user_id: int,
    payload: UpdateUserProfileRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    new_name = _clean_optional_str(payload.name)
    new_last = _clean_optional_str(payload.last_name)
    new_phone = _clean_optional_str(payload.phone)
    new_role = _clean_optional_str(payload.role)
    new_company_name = _clean_optional_str(payload.company_name)
    new_cnpj = _clean_optional_str(payload.cnpj)
    new_segment = _clean_optional_str(payload.segment)

    if new_name is not None:
        user.name = new_name
    if new_last is not None:
        user.last_name = new_last
    if new_phone is not None:
        user.phone = new_phone
    if new_role is not None:
        user.role = new_role

    if user.created_at is None:
        user.created_at = datetime.utcnow()

    company = db.get(Company, user.company_id) if user.company_id else None
    has_company_data = any(
        v is not None for v in (new_company_name, new_cnpj, new_segment)
    )
    if has_company_data:
        if company is None:
            company = Company(
                name=new_company_name or "",
                cnpj=new_cnpj,
                segment=new_segment,
                status="ACTIVE",
                created_at=datetime.utcnow(),
            )
            db.add(company)
            db.flush()
            user.company_id = company.id
        else:
            if new_company_name is not None:
                company.name = new_company_name
            if new_cnpj is not None:
                company.cnpj = new_cnpj
            if new_segment is not None:
                company.segment = new_segment
            company.updated_at = datetime.utcnow()
            db.add(company)

    db.add(user)
    db.commit()
    db.refresh(user)
    company = db.get(Company, user.company_id) if user.company_id else None

    return {
        "ok": True,
        "user": {
            "id": user.id,
            "name": user.name,
            "last_name": user.last_name,
            "email": user.email,
            "role": user.role or "user",
            "phone": user.phone,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "company": {
                "id": company.id,
                "name": company.name,
                "cnpj": company.cnpj,
                "segment": company.segment,
                "status": company.status,
            } if company else None,
        },
    }


@app.post("/assessment/save")
def save_assessment(
    payload: SaveAssessmentRequest,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    saved = 0
    for item in payload.answers:
        row = db.scalar(
            select(Answer).where(
                Answer.assessment_id == payload.assessment_id,
                Answer.question_id == item.question_id,
            )
        )
        if not row:
            row = Answer(
                assessment_id=payload.assessment_id,
                question_id=item.question_id,
                created_at=datetime.utcnow(),
            )
        row.score = item.score
        row.justification = item.justification or ""
        row.evidence = item.evidence or ""
        row.updated_at = datetime.utcnow()
        db.add(row)
        saved += 1

    db.flush()

    assessment = db.get(Assessment, payload.assessment_id)
    if assessment is not None:
        total_questions, _ = _question_catalog_stats(db)
        answered_total = len(
            db.scalars(
                select(Answer.id).where(Answer.assessment_id == payload.assessment_id)
            ).all()
        )
        if total_questions > 0 and answered_total >= total_questions:
            assessment.status = "COMPLETED"
            assessment.progress_percent = 100.0
        else:
            assessment.status = "IN_PROGRESS"
            assessment.progress_percent = round(
                (answered_total / total_questions) * 100, 2
                if total_questions > 0
                else 0.0,
            )
        assessment.updated_at = datetime.utcnow()
        db.add(assessment)

    db.commit()
    return {"ok": True, "saved": saved}
