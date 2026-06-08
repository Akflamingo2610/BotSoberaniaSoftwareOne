"""Carrega o catálogo oficial de perguntas a partir do Excel Assessment_OTIMIZADO."""

from __future__ import annotations

from pathlib import Path

ORIGEM_TO_PHASE = {
    "original": "Quick_Wins",
    "lens soberania": "Foundational",
    "sec assessment": "Efficient",
}

PHASE_ORDER = ["Quick_Wins", "Foundational", "Efficient", "Optimized"]
PILAR_ORDER = ["Compliance", "Continuity", "Control"]

DEFAULT_XLSX_NAME = "Assessment_OTIMIZADO_176_Perguntas_v4 1.xlsx"


def _norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_ok(value: object) -> bool:
    return _norm(value).lower() == "ok"


def _phase_from_origem(origem: str) -> str:
    key = _norm(origem).lower()
    if key not in ORIGEM_TO_PHASE:
        raise ValueError(
            f"Origem desconhecida: {origem!r} (esperado uma de {list(ORIGEM_TO_PHASE)})"
        )
    return ORIGEM_TO_PHASE[key]


def _sort_key(row: dict) -> tuple[int, int, str]:
    phase = row["phase"]
    pilar = row["pilar"]
    code = row["question_code"] or ""
    phase_idx = PHASE_ORDER.index(phase) if phase in PHASE_ORDER else 99
    pilar_idx = PILAR_ORDER.index(pilar) if pilar in PILAR_ORDER else 99
    return (phase_idx, pilar_idx, code)


def load_rows_from_excel(path: Path) -> list[dict]:
    """Importa linhas com Visualizar Soberania = ok (24 perguntas por pilar)."""
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl nao instalado") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Assessment" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Planilha 'Assessment' nao encontrada. Abas: {wb.sheetnames}")
    ws = wb["Assessment"]

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        if not row or len(row) < 3:
            continue
        header = _norm(row[2]).lower()
        if "visualizar" in header and "soberania" in header:
            header_row_idx = i
            break
    if header_row_idx is None:
        wb.close()
        raise ValueError("Cabecalho com coluna 'Visualizar Soberania' nao encontrado.")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[1] is None:
            continue
        if not _cell_ok(row[2]):
            continue
        code = _norm(row[1])
        pilar = _norm(row[4])
        origem = _norm(row[5])
        pergunta = _norm(row[6])
        dominio = _norm(row[7]) or None
        pilar_tec = _norm(row[8]) or None
        aws = _norm(row[9]) or None
        evid = _norm(row[11]) or None
        if not pergunta:
            continue
        if pilar not in PILAR_ORDER:
            raise ValueError(f"Pilar invalido na linha {code}: {pilar!r}")
        rows.append(
            {
                "question_code": code,
                "phase": _phase_from_origem(origem),
                "pilar": pilar,
                "recommendation": pergunta,
                "dominio": dominio,
                "pilar_tecnico": pilar_tec,
                "aws_service": aws,
                "norms": evid,
            }
        )
    wb.close()
    rows.sort(key=_sort_key)
    return rows
