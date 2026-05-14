#!/usr/bin/env python3
"""
Importa perguntas do Excel Assessment_OTIMIZADO_* para a tabela `questions`.

Regras:
- Importa somente linhas com coluna "Visualizar Soberania" = ok (case-insensitive).
- Pilar: coluna "Pilar 3Cs" (Compliance | Continuity | Control).
- Fase (coluna Origem do Excel -> phase do banco, alinhado ao app Flutter):
    Original        -> Quick_Wins
    Lens Soberania  -> Foundational
    Sec Assessment  -> Efficient

ATENCAO: por padrao este script APAGA todas as respostas (`answers`) e todas as
perguntas (`questions`) antes de inserir, para evitar FK quebrada e duplicata.
Use --dry-run para validar sem gravar.

Uso (na pasta python_api, com as mesmas variaveis de ambiente da API, ex. .env.dev):

  python scripts/import_assessment_excel.py --dry-run
  python scripts/import_assessment_excel.py --apply

Opcional:

  python scripts/import_assessment_excel.py --apply --xlsx "C:\\caminho\\arquivo.xlsx"
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

# Permite rodar sem instalar o pacote app
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from sqlalchemy import delete, func, select

from app.models import Answer, Question


DEFAULT_XLSX = _ROOT.parent / "Assessment_OTIMIZADO_176_Perguntas_v4 1.xlsx"

ORIGEM_TO_PHASE = {
    "original": "Quick_Wins",
    "lens soberania": "Foundational",
    "sec assessment": "Efficient",
}

PHASE_ORDER = ["Quick_Wins", "Foundational", "Efficient", "Optimized"]
PILAR_ORDER = ["Compliance", "Continuity", "Control"]


def _norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _cell_ok(val: object) -> bool:
    s = _norm(val).lower()
    return s == "ok"


def _phase_from_origem(origem: str) -> str:
    key = _norm(origem).lower()
    if key not in ORIGEM_TO_PHASE:
        raise ValueError(f"Origem desconhecida: {origem!r} (esperado uma de {list(ORIGEM_TO_PHASE)})")
    return ORIGEM_TO_PHASE[key]


def _sort_key(row: dict) -> tuple[int, int, str]:
    phase = row["phase"]
    pilar = row["pilar"]
    code = row["question_code"] or ""
    pi = PHASE_ORDER.index(phase) if phase in PHASE_ORDER else 99
    qi = PILAR_ORDER.index(pilar) if pilar in PILAR_ORDER else 99
    return (pi, qi, code)


def load_rows_from_excel(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("Instale openpyxl: pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Assessment" not in wb.sheetnames:
        raise SystemExit(f"Planilha 'Assessment' nao encontrada. Abas: {wb.sheetnames}")
    ws = wb["Assessment"]

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        if not row or len(row) < 3:
            continue
        h2 = _norm(row[2]).lower()
        if "visualizar" in h2 and "soberania" in h2:
            header_row_idx = i
            break
    if header_row_idx is None:
        wb.close()
        raise SystemExit("Cabecalho com coluna 'Codigo' nao encontrado.")

    out: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[1] is None:
            continue
        if not _cell_ok(row[2]):
            continue
        code = _norm(row[1])
        pilar = _norm(row[4])
        origem = _norm(row[5])
        pergunta = _norm(row[6])
        dominio = _norm(row[7]) or None
        pilar_tec = _norm(row[8]) or None
        aws = _norm(row[9]) or None
        evid = _norm(row[11]) or None
        if not pergunta:
            continue
        if pilar not in PILAR_ORDER:
            raise ValueError(f"Pilar invalido na linha {code}: {pilar!r}")
        phase = _phase_from_origem(origem)
        out.append(
            {
                "question_code": code,
                "phase": phase,
                "pilar": pilar,
                "recommendation": pergunta,
                "dominio": dominio,
                "pilar_tecnico": pilar_tec,
                "aws_service": aws,
                "norms": evid,
            }
        )
    wb.close()
    out.sort(key=_sort_key)
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="Importa perguntas do Excel (Visualizar Soberania=ok).")
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="Caminho do .xlsx",
    )
    ap.add_argument("--dry-run", action="store_true", help="So le e imprime resumo, sem gravar.")
    ap.add_argument(
        "--apply",
        action="store_true",
        help="Apaga answers+questions e insere (DESTRUTIVO).",
    )
    args = ap.parse_args()

    if not args.dry_run and not args.apply:
        ap.error("Informe --dry-run ou --apply")

    path: Path = args.xlsx
    if not path.is_file():
        raise SystemExit(f"Arquivo nao encontrado: {path}")

    rows = load_rows_from_excel(path)
    print(f"Lidas {len(rows)} perguntas (Visualizar Soberania=ok).")
    from collections import Counter

    c_phase = Counter(r["phase"] for r in rows)
    c_pilar = Counter(r["pilar"] for r in rows)
    print("Por fase:", dict(c_phase))
    print("Por pilar:", dict(c_pilar))

    if args.dry_run:
        print("Dry-run: nenhuma alteracao no banco.")
        for r in rows[:3]:
            print("  exemplo:", r["question_code"], r["phase"], r["pilar"], r["recommendation"][:60] + "...")
        return

    from app.database import SessionLocal

    now = datetime.utcnow()
    with SessionLocal() as db:
        n_ans = int(db.scalar(select(func.count()).select_from(Answer)) or 0)
        n_q = int(db.scalar(select(func.count()).select_from(Question)) or 0)
        print(f"Antes: {n_ans} respostas, {n_q} perguntas. Apagando...")
        db.execute(delete(Answer))
        db.execute(delete(Question))
        db.flush()
        for i, r in enumerate(rows, start=1):
            db.add(
                Question(
                    created_at=now,
                    phase=r["phase"],
                    pilar=r["pilar"],
                    recommendation=r["recommendation"],
                    weight=None,
                    order_index=i,
                    question_code=r["question_code"],
                    dominio=r["dominio"],
                    recommendation_en=None,
                    recommendation_es=None,
                    guidance=None,
                    how_to_check=None,
                    aws_service=r["aws_service"],
                    pilar_tecnico=r["pilar_tecnico"],
                    norms=r["norms"],
                )
            )
        db.commit()
        nq2 = int(db.scalar(select(func.count()).select_from(Question)) or 0)
        print(f"OK: inseridas {nq2} perguntas (respostas anteriores removidas).")


if __name__ == "__main__":
    main()
